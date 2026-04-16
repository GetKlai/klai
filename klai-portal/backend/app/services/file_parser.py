"""Parse uploaded files into plain text for knowledge ingestion."""

import io

import structlog

logger = structlog.get_logger()


def parse_file(content: bytes, filename: str) -> str:
    """Parse file content to plain text. Supports PDF, DOCX, XLSX, TXT, MD, CSV, JSON, XML."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "pdf":
        return _parse_pdf(content)
    elif ext == "docx":
        return _parse_docx(content)
    elif ext in ("xlsx", "xls"):
        return _parse_xlsx(content)
    elif ext in ("txt", "md", "csv", "json", "xml", "html", "htm", "rst"):
        return content.decode("utf-8", errors="replace")
    else:
        # Try UTF-8 as fallback
        return content.decode("utf-8", errors="replace")


def _parse_pdf(content: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)
    return "\n\n".join(pages)


def _parse_docx(content: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _parse_xlsx(content: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            sheets.append(f"## {ws.title}\n" + "\n".join(rows))
    return "\n\n".join(sheets)


def parse_image_ocr(content: bytes, filename: str) -> str:
    """Extract text from image using OCR. Returns empty string if Tesseract not available."""
    try:
        import pytesseract
        from PIL import Image

        img = Image.open(io.BytesIO(content))
        text = pytesseract.image_to_string(img)
        return text.strip()
    except Exception:
        logger.warning("OCR failed or Tesseract not available", filename=filename)
        # Fallback: store as image reference
        return f"[Image: {filename}]"
